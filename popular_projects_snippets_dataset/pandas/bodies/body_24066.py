# Extracted from ./data/repos/pandas/pandas/io/excel/_openpyxl.py
# Use the openpyxl module as the Excel writer.
from openpyxl.workbook import Workbook

engine_kwargs = combine_kwargs(engine_kwargs, kwargs)

super().__init__(
    path,
    mode=mode,
    storage_options=storage_options,
    if_sheet_exists=if_sheet_exists,
    engine_kwargs=engine_kwargs,
)

# ExcelWriter replaced "a" by "r+" to allow us to first read the excel file from
# the file and later write to it
if "r+" in self._mode:  # Load from existing workbook
    from openpyxl import load_workbook

    self._book = load_workbook(self._handles.handle, **engine_kwargs)
    self._handles.handle.seek(0)
else:
    # Create workbook object with default optimized_write=True.
    self._book = Workbook(**engine_kwargs)

    if self.book.worksheets:
        self.book.remove(self.book.worksheets[0])
